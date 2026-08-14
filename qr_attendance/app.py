from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time

app = Flask(__name__)

app.secret_key = "attendance-secret-key"

EXCEL_FILE = "attendance.xlsx"


def create_excel_file():
    """Create the Excel file if it does not exist."""

    if not os.path.exists(EXCEL_FILE):

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Attendance"

        sheet.append([
            "No",
            "Adı Soyadı",
            "Fakülte",
            "Bölüm",
            "Katılım Durumu",
            "Tarih",
            "Saat"
        ])

        workbook.save(EXCEL_FILE)
        workbook.close()


def add_attendance(name, faculty, department):
    """Add an attendance record to Excel."""

    # Try several times if Excel is temporarily locked
    for attempt in range(5):

        try:

            workbook = load_workbook(EXCEL_FILE)

            sheet = workbook["Attendance"]

            # Check for duplicate registration
            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                existing_name = row[1]
                existing_faculty = row[2]
                existing_department = row[3]

                if (
                    existing_name
                    and existing_faculty
                    and existing_department
                    and str(existing_name).strip().lower()
                    == name.strip().lower()
                    and str(existing_faculty).strip().lower()
                    == faculty.strip().lower()
                    and str(existing_department).strip().lower()
                    == department.strip().lower()
                ):

                    workbook.close()

                    return False

            # Create registration number
            registration_number = sheet.max_row

            # Current date and time
            now = datetime.now()

            sheet.append([
                registration_number,
                name,
                faculty,
                department,
                "KATILDI",
                now.strftime("%d.%m.%Y"),
                now.strftime("%H:%M:%S")
            ])

            workbook.save(EXCEL_FILE)

            workbook.close()

            return True

        except PermissionError:

            # Excel file may temporarily be locked
            if attempt < 4:

                time.sleep(1)

            else:

                return "locked"

        except Exception as error:

            print("Excel error:", error)

            return "error"


@app.route("/", methods=["GET", "POST"])
def index():

    if request.method == "POST":

        name = request.form.get(
            "name",
            ""
        ).strip()

        faculty = request.form.get(
            "faculty",
            ""
        ).strip()

        department = request.form.get(
            "department",
            ""
        ).strip()

        # Make sure all fields are filled
        if not name or not faculty or not department:

            flash(
                "Lütfen tüm alanları doldurunuz.",
                "error"
            )

            return redirect(
                url_for("index")
            )

        # Save attendance
        result = add_attendance(
            name,
            faculty,
            department
        )

        if result is True:

            return render_template(
                "success.html",
                name=name
            )

        elif result is False:

            flash(
                "Bu kişi daha önce katılım kaydı oluşturmuştur.",
                "error"
            )

            return redirect(
                url_for("index")
            )

        elif result == "locked":

            flash(
                "Kayıt dosyası şu anda kullanımda. "
                "Lütfen Excel dosyasını kapatıp tekrar deneyiniz.",
                "error"
            )

            return redirect(
                url_for("index")
            )

        else:

            flash(
                "Kayıt sırasında bir hata oluştu. "
                "Lütfen tekrar deneyiniz.",
                "error"
            )

            return redirect(
                url_for("index")
            )

    return render_template(
        "index.html"
    )


if __name__ == "__main__":

    create_excel_file()

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )